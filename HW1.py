import sqlite3
from openpyxl import Workbook, load_workbook
from PyQt6.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QFileDialog, QMessageBox

# Путь к базе данных SQLite
DATABASE_PATH = 'database.db'

def import_data():
    file_path, _ = QFileDialog.getOpenFileName(None, "Выберите Excel файл", "", "Excel Files (*.xlsx)")
    if file_path:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            connection = sqlite3.connect(DATABASE_PATH)
            cursor = connection.cursor()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, quantity, price = row
                cursor.execute("INSERT INTO products (name, quantity, price) VALUES (?, ?, ?)", (name, quantity, price))

            connection.commit()
            connection.close()

            show_message_box("Импорт данных", "Данные успешно импортированы из Excel файла.")
        except Exception as e:
            show_message_box("Ошибка", f"Произошла ошибка при импорте данных:\n{str(e)}")

def export_data():
    file_path, _ = QFileDialog.getSaveFileName(None, "Сохранить как Excel файл", "", "Excel Files (*.xlsx)")
    if file_path:
        try:
            connection = sqlite3.connect(DATABASE_PATH)
            cursor = connection.cursor()
            cursor.execute("SELECT id, name, quantity, price FROM products")
            rows = cursor.fetchall()

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "Наименование", "Количество", "Цена"])

            for row in rows:
                sheet.append(row)

            workbook.save(file_path)
            connection.close()

            show_message_box("Экспорт данных", "Данные успешно экспортированы в Excel файл.")
        except Exception as e:
            show_message_box("Ошибка", f"Произошла ошибка при экспорте данных:\n{str(e)}")

def show_message_box(title, message):
    message_box = QMessageBox()
    message_box.setWindowTitle(title)
    message_box.setText(message)
    message_box.exec()

app = QApplication([])
window = QMainWindow()

widget = QWidget()
layout = QVBoxLayout()

button_import = QPushButton("Импортировать")
button_import.clicked.connect(import_data)
layout.addWidget(button_import)

button_export = QPushButton("Экспортировать")
button_export.clicked.connect(export_data)
layout.addWidget(button_export)

widget.setLayout(layout)
window.setCentralWidget(widget)
window.show()

app.exec()
